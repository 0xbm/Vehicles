import datetime

from openpyxl import Workbook
import calendar
import openpyxl
import datetime

# class Style:
#     def __init__(self, Workbook, ws):
#         self.Workbook = Workbook()
#         self.ws = ws.active

#     def months(self, wb=Workbook):

#         months = [calendar.month_name[i] for i in range(1, 12)]
#         print(months)

#         for worksheet in months:
#             wb.create_sheet(worksheet)

#         wb.save("style.xlsx")

#     # def merge_cells(ws):
#     #     ws.merge_cells("A1:E1")
#     #     ws.merge_cells("A3:C3")
#     #     ws.merge_cells("A4:C4")
#     #     ws.merge_cells("A1:B1")
#     #     ws.merge_cells("B7:D7")
#     #     ws.merge_cells("B8:D8")
#     #     ws.merge_cells("B9:D9")
#     #     ws.merge_cells("B10:D10")
#     #     ws.merge_cells("B11:D11")

#     def save_style(self, wb):
#         create_workbook.wb.save("style.xlsx")


# style = Style
# style.months(Workbook)


"""
wb = Workbook()
ws = wb.active
 months = [calendar.month_name[i] for i in range(1, 12)]
        for worksheet in months:
            wb.create_sheet(worksheet)
del wb["Sheet"]

wb.save("style.xlsx")

"""


# wb = Workbook()
# # ws = wb.active

# months = [calendar.month_name[i] for i in range(1, 13)]
# for worksheet in months:
#     wb.create_sheet(worksheet)
# del wb["Sheet"]

# ws1 = wb[months[0]]
# ws1["A1"] = 56
# ws1.merge_cells("A1:E1")
# wb.save("style.xlsx")
#######################
class Style:
    def create_months(self, wb=Workbook()):
        # wb = Workbook()
        # ws = wb.active
        months = [calendar.month_name[i] for i in range(1, 13)]

        for worksheet in months:
            wb.create_sheet(worksheet)

        del wb["Sheet"]
        wb.save("style.xlsx")

    def merge_cells(self):
        # wb = Workbook()
        wb = openpyxl.load_workbook("style.xlsx")
        # ws = wb[months[0]]
        # ws = wb.get_sheet_by_name["January"]
        ws = wb["January"]
        #  ustaw zmiena sheet np na luty.
        ws.merge_cells("A1:E1")
        ws.merge_cells("A3:C3")
        ws.merge_cells("A4:B4")
        ws.merge_cells("A1:B1")
        ws.merge_cells("A6:C6")
        ws.merge_cells("A7:C7")
        ws.merge_cells("A8:C8")
        ws.merge_cells("A9:C9")
        ws.merge_cells("A10:C10")
        ws.merge_cells("A11:C11")

        wb.save("style.xlsx")

        # def copy(self):
        # wb = openpyxl.load_workbook("style.xlsx")
        # ws = wb["January"]

        # def sheetnames(self):
        # wb = openpyxl.load_workbook("style.xlsx")
        # for sheet in wb.worksheets:
        # print(wb.sheetnames)


class Text:
    def card_name(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A1'] = 'ROZLICZENIE MIESIĘCZNE ZUŻYCIA PALIWA'
        wb.save("style.xlsx")

    def brand(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A3'] = 'Pojazd służbowy, marka'
        wb.save("style.xlsx")

    def registration_number(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A4'] = 'Nr rejestracyjny'
        wb.save("style.xlsx")

    def month(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A5'] = 'Miesiąc'
        ws['B5'] = wb.sheetnames[0]
        wb.save("style.xlsx")

    def year(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['C5'] = 'rok'
        year = datetime.datetime.today().year

        ws['D5'] = year
        wb.save("style.xlsx")

    def count_beginning_month(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A6'] = '1. Stan licznika na początku miesiąca'
        wb.save("style.xlsx")

    def count_ending_month(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A7'] = '2. Stan licznika na końcu miesiąca'
        wb.save("style.xlsx")

    def km_in_month(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A8'] = '3. Przejechano km/mth w miesiącu'
        wb.save("style.xlsx")

    def fuel__reamaining_last_month(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A9'] = '4.	Stan paliwa pozostałego z ubiegłego miesiąca'
        wb.save("style.xlsx")

    def fuel_purchased_in_month(self):
        wb = openpyxl.load_workbook("style.xlsx")
        ws = wb.active

        ws['A10'] = '5.Ilość paliwa zakupionego w miesiącu '
        wb.save("style.xlsx")


styl = Style()
styl.create_months()
styl.merge_cells()

text = Text()
text.card_name()
text.brand()
text.registration_number()
text.month()
text.year()
text.count_beginning_month()
text.count_ending_month()
text.fuel__reamaining_last_month()
text.fuel_purchased_in_month()
text.km_in_month()

# styl.sheetnames()

# tod zrob wzor arkusza w openpyxl
# tod zrob kopiowanie arkusza na kolejne miesiace

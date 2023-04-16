import openpyxl
import datetime
from openpyxl.styles import Font

wb = openpyxl.load_workbook("card.xlsx")
ws = wb.active


class Text:
    def card_name(self):
        ws['A1'] = 'ROZLICZENIE MIESIĘCZNE ZUŻYCIA PALIWA'
        ws['A1'].font =Font(bold=True)
    def brand(self):
        ws['A3'] = 'Pojazd służbowy, marka'

    def registration_number(self):
        ws['A4'] = 'Nr rejestracyjny'

    def month(self):
        ws['A5'] = 'Miesiąc'
        ws['B5'] = wb.sheetnames[0][0:3]
        ws['B5'].font = Font(bold=True)
    def year(self):
        ws['C5'] = 'rok'
        year = datetime.datetime.today().year

        ws['D5'] = year
        ws['D5'].font =Font(bold=True)

    def count_beginning_month(self):
        ws['A6'] = '1. Stan licznika na początku miesiąca'

    def count_ending_month(self):
        ws['A7'] = '2. Stan licznika na końcu miesiąca'

    def km_in_month(self):
        ws['A8'] = '3. Przejechano km/mth w miesiącu'

    def fuel__reamaining_last_month(self):
        ws['A9'] = '4.	Stan paliwa pozostałego z ubiegłego miesiąca'

    def fuel_purchased_in_month(self):
        ws['A10'] = '5.Ilość paliwa zakupionego w miesiącu'

    def total_fuel(self):
        ws['A24'] = '6.Razem paliwa'

    def fuel_consumption(self):
        ws['A25'] = '7.Faktyczne zyżycie paliwa'

    def limit_norm(self):
        ws['A26'] = '8.Norma graniczna'

    def excessive_consumption(self):
        ws['A27'] = '9.Zużycie ponadnormatywne %'

    def fuel_for_next_month(self):
        ws['A28'] = '10.Pozostało paliwa na m-c następny'

    def table_text(self):
        ws['B12'] = 'LP'
        ws['C12'] = 'Data zakupu'
        ws['D12'] = 'Olej napędowy (L)'
        ws['E12'] = 'Benzyna (L)'
        ws['B31'] = 'Podpis'
        ws['D30'] = 'Sporządził'
        ws['E30'] = 'Sprawdził'
        ws['B32'] = 'Data'


text = Text()
text.card_name()
text.brand()
text.registration_number()
text.month()
text.year()
text.count_beginning_month()
text.count_ending_month()
text.fuel__reamaining_last_month()
text.fuel_purchased_in_month()
text.km_in_month()
text.total_fuel()
text.fuel_consumption()
text.limit_norm()
text.excessive_consumption()
text.fuel_for_next_month()
text.table_text()
wb.save("card.xlsx")

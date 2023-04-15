import openpyxl
from openpyxl.styles import borders
from openpyxl.styles.borders import Border

wb = openpyxl.load_workbook("card.xlsx")
ws = wb.active


class CreateTable:
    def tank_table(self):
        border_thin = borders.Side(style=None, color='FF000000', border_style='thin')
        border0 = borders.Side(style=None, color=None, border_style=None)
        thin = Border(left=border_thin, right=border_thin, bottom=border_thin, top=border_thin)

        for row in ws.iter_rows(min_row=12, min_col=2, max_row=22, max_col=5):
            for cell in row:
                cell.border = thin
    def signatures_table(self):
        border_thin = borders.Side(style=None, color='FF000000', border_style='thin')
        border0 = borders.Side(style=None, color=None, border_style=None)
        thin = Border(left=border_thin, right=border_thin, bottom=border_thin, top=border_thin)

        for row in ws.iter_rows(min_row=30, min_col=2, max_row=32, max_col=5):
            for cell in row:
                cell.border = thin

    #wb.save('card.xlsx')


create_table = CreateTable()
create_table.tank_table()
create_table.signatures_table()
wb.save('card.xlsx')

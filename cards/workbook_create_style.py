import calendar
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook


class Style:
    def create_months(self, wb=Workbook()):
        months = [calendar.month_name[i] for i in range(1, 13)]

        for worksheet in months:
            wb.create_sheet(worksheet)

        del wb["Sheet"]
        wb.save("card.xlsx")

    def merge_cells(self):
        wb = openpyxl.load_workbook("card.xlsx")
        ws = wb["January"]
        #  ustaw zmiena sheet np na luty.
        ws.merge_cells("A1:E1")
        ws.merge_cells("A3:C3")
        ws.merge_cells("A4:B4")
        ws.merge_cells("A1:B1")
        ws.merge_cells("A6:D6")
        ws.merge_cells("A7:D7")
        ws.merge_cells("A8:D8")
        ws.merge_cells("A9:D9")
        ws.merge_cells("A10:D10")
        ws.merge_cells("A24:D24")
        ws.merge_cells("A25:D25")
        ws.merge_cells("A26:D26")
        ws.merge_cells("A27:D27")
        ws.merge_cells("A28:D28")
        ws.merge_cells("B30:C30")
        ws.merge_cells("B31:C31")
        ws.merge_cells("B32:C32")

        wb.save("card.xlsx")
    def adjust_rows_cols_dimension(self):
        wb = openpyxl.load_workbook("card.xlsx")
        ws = wb["January"]

        ws.row_dimensions[31].height = 30
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        wb.save("card.xlsx")
    def aligment(self):
        wb = openpyxl.load_workbook("card.xlsx")
        ws = wb["January"]
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["C5"].alignment = Alignment(horizontal="right")
        ws["D5"].alignment = Alignment(horizontal="left")
        ws["B12"].alignment = Alignment(horizontal="center")
        ws["C12"].alignment = Alignment(horizontal="center")
        ws["D12"].alignment = Alignment(horizontal="center")
        ws["E12"].alignment = Alignment(horizontal="center")

        wb.save("card.xlsx")


template = Style()
template.create_months()
template.merge_cells()
template.adjust_rows_cols_dimension()
template.aligment()
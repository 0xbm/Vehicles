from openpyxl import Workbook
import calendar


# class Style:
#     def __init__(self, Workbook, ws):
#         self.Workbook = Workbook()
#         self.ws = ws.active

#     def months(self, wb=Workbook):

#         months = [calendar.month_name[i] for i in range(1, 12)]
#         print(months)

#         for worksheet in months:
#             wb.create_sheet(worksheet)

#         wb.save("style.xlsx")

#     # def merge_cells(ws):
#     #     ws.merge_cells("A1:E1")
#     #     ws.merge_cells("A3:C3")
#     #     ws.merge_cells("A4:C4")
#     #     ws.merge_cells("A1:B1")
#     #     ws.merge_cells("B7:D7")
#     #     ws.merge_cells("B8:D8")
#     #     ws.merge_cells("B9:D9")
#     #     ws.merge_cells("B10:D10")
#     #     ws.merge_cells("B11:D11")

#     def save_style(self, wb):
#         create_workbook.wb.save("style.xlsx")


# style = Style
# style.months(Workbook)


"""
wb = Workbook()
ws = wb.active
 months = [calendar.month_name[i] for i in range(1, 12)]
        for worksheet in months:
            wb.create_sheet(worksheet)
del wb["Sheet"]

wb.save("style.xlsx")

"""
# wb = Workbook()
# # ws = wb.active

# months = [calendar.month_name[i] for i in range(1, 13)]
# for worksheet in months:
#     wb.create_sheet(worksheet)
# del wb["Sheet"]

# ws1 = wb[months[0]]
# ws1["A1"] = 56
# ws1.merge_cells("A1:E1")
# wb.save("style.xlsx")
#######################
class Style:
    def months(self, wb=Workbook()):
        # wb = Workbook()
        # ws = wb.active
        months = [calendar.month_name[i] for i in range(1, 13)]

        for worksheet in months:
            wb.create_sheet(worksheet)
        del wb["Sheet"]
        wb.save("style.xlsx")

    def merge_cells(self):
        wb = Workbook()
        # ws = wb[months[0]]
        ws = wb.get_sheet_by_name["January"]
        # usatw zmiene sheet np na luty
        ws.merge_cells("A1:E1")
        ws.merge_cells("A3:C3")
        ws.merge_cells("A4:C4")
        ws.merge_cells("A1:B1")
        ws.merge_cells("B7:D7")
        ws.merge_cells("B8:D8")
        ws.merge_cells("B9:D9")
        ws.merge_cells("B10:D10")
        ws.merge_cells("B11:D11")


styl = Style()
# styl.months()
styl.merge_cells()
